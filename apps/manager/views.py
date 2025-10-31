# Importa o modelo LogEntry da biblioteca django-auditlog
import io
import unicodedata
from collections import defaultdict
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal

import openpyxl
from auditlog.models import LogEntry
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from xhtml2pdf import pisa

from apps.patients.models import ClinicalWarningSign, ConsultationRecord, Patient, Record


def normalize_str(s):
    """
    Remove acentos, converte para minúsculo, remove espaços extras
    e aplica title case. Ex: ' MãCeIó ' -> 'Maceio'
    """
    if not s:
        return ""
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("utf-8").strip().title()


MICRORREGIAO_POR_MUNICIPIO = {
    "Maceió": "1ª Região de Saúde",
    "Barra de Santo Antônio": "1ª Região de Saúde",
    "Barra de São Miguel": "1ª Região de Saúde",
    "Coqueiro Seco": "1ª Região de Saúde",
    "Flexeiras": "1ª Região de Saúde",
    "Marechal Deodoro": "1ª Região de Saúde",
    "Messias": "1ª Região de Saúde",
    "Paripueira": "1ª Região de Saúde",
    "Pilar": "1ª Região de Saúde",
    "Rio Largo": "1ª Região de Saúde",
    "Santa Luzia do Norte": "1ª Região de Saúde",
    "Satuba": "1ª Região de Saúde",
    "Jacuípe": "2ª Região de Saúde",
    "Japaratinga": "2ª Região de Saúde",
    "Maragogi": "2ª Região de Saúde",
    "Matriz de Camaragibe": "2ª Região de Saúde",
    "Passo de Camaragibe": "2ª Região de Saúde",
    "Porto Calvo": "2ª Região de Saúde",
    "Porto de Pedras": "2ª Região de Saúde",
    "São Luís do Quitunde": "2ª Região de Saúde",
    "São Miguel dos Milagres": "2ª Região de Saúde",
    "Branquinha": "3ª Região de Saúde",
    "Campestre": "3ª Região de Saúde",
    "Colônia Leopoldina": "3ª Região de Saúde",
    "Ibateguara": "3ª Região de Saúde",
    "Joaquim Gomes": "3ª Região de Saúde",
    "Jundiá": "3ª Região de Saúde",
    "Novo Lino": "3ª Região de Saúde",
    "Santana do Mundaú": "3ª Região de Saúde",
    "São José da Laje": "3ª Região de Saúde",
    "União dos Palmares": "3ª Região de Saúde",
    "Atalaia": "4ª Região de Saúde",
    "Cajueiro": "4ª Região de Saúde",
    "Capela": "4ª Região de Saúde",
    "Chã Preta": "4ª Região de Saúde",
    "Mar Vermelho": "4ª Região de Saúde",
    "Maribondo": "4ª Região de Saúde",
    "Murici": "4ª Região de Saúde",
    "Paulo Jacinto": "4ª Região de Saúde",
    "Pindoba": "4ª Região de Saúde",
    "Quebrangulo": "4ª Região de Saúde",
    "Viçosa": "4ª Região de Saúde",
    "Anadia": "5ª Região de Saúde",
    "Boca da Mata": "5ª Região de Saúde",
    "Campo Alegre": "5ª Região de Saúde",
    "Junqueiro": "5ª Região de Saúde",
    "Limoeiro de Anadia": "5ª Região de Saúde",
    "Roteiro": "5ª Região de Saúde",
    "São Miguel dos Campos": "5ª Região de Saúde",
    "Tanque d'Arca": "5ª Região de Saúde",
    "Teotônio Vilela": "5ª Região de Saúde",
    "Coruripe": "6ª Região de Saúde",
    "Feliz Deserto": "6ª Região de Saúde",
    "Igreja Nova": "6ª Região de Saúde",
    "Jequiá da Praia": "6ª Região de Saúde",
    "Olho d'Água Grande": "6ª Região de Saúde",
    "Penedo": "6ª Região de Saúde",
    "Piaçabuçu": "6ª Região de Saúde",
    "Porto Real do Colégio": "6ª Região de Saúde",
    "São Brás": "6ª Região de Saúde",
    "Arapiraca": "7ª Região de Saúde",
    "Batalha": "7ª Região de Saúde",
    "Belo Monte": "7ª Região de Saúde",
    "Campo Grande": "7ª Região de Saúde",
    "Coité do Nóia": "7ª Região de Saúde",
    "Craíbas": "7ª Região de Saúde",
    "Feira Grande": "7ª Região de Saúde",
    "Girau do Ponciano": "7ª Região de Saúde",
    "Jacaré dos Homens": "7ª Região de Saúde",
    "Jaramataia": "7ª Região de Saúde",
    "Lagoa da Canoa": "7ª Região de Saúde",
    "Monteirópolis": "7ª Região de Saúde",
    "Olho d'Água das Flores": "7ª Região de Saúde",
    "Olivença": "7ª Região de Saúde",
    "São Sebastião": "7ª Região de Saúde",
    "Taquarana": "7ª Região de Saúde",
    "Traipu": "7ª Região de Saúde",
    "Belém": "8ª Região de Saúde",
    "Cacimbinhas": "8ª Região de Saúde",
    "Estrela de Alagoas": "8ª Região de Saúde",
    "Igaci": "8ª Região de Saúde",
    "Major Isidoro": "8ª Região de Saúde",
    "Minador do Negrão": "8ª Região de Saúde",
    "Palmeira dos Índios": "8ª Região de Saúde",
    "Canapi": "9ª Região de Saúde",
    "Carneiros": "9ª Região de Saúde",
    "Dois Riachos": "9ª Região de Saúde",
    "Maravilha": "9ª Região de Saúde",
    "Ouro Branco": "9ª Região de Saúde",
    "Palestina": "9ª Região de Saúde",
    "Pão de Açúcar": "9ª Região de Saúde",
    "Poço das Trincheiras": "9ª Região de Saúde",
    "Santana do Ipanema": "9ª Região de Saúde",
    "São José da Tapera": "9ª Região de Saúde",
    "Senador Rui Palmeira": "9ª Região de Saúde",
    "Água Branca": "10ª Região de Saúde",
    "Delmiro Gouveia": "10ª Região de Saúde",
    "Inhapi": "10ª Região de Saúde",
    "Mata Grande": "10ª Região de Saúde",
    "Olho d'Água do Casado": "10ª Região de Saúde",
    "Pariconha": "10ª Região de Saúde",
    "Piranhas": "10ª Região de Saúde",
}

MICRORREGIAO_POR_MUNICIPIO = {normalize_str(k): v for k, v in MICRORREGIAO_POR_MUNICIPIO.items()}

MACRORREGIAO_POR_MICRORREGIAO = {
    "1ª Região de Saúde": "Macrorregião I",
    "2ª Região de Saúde": "Macrorregião I",
    "3ª Região de Saúde": "Macrorregião I",
    "4ª Região de Saúde": "Macrorregião I",
    "5ª Região de Saúde": "Macrorregião I",
    "10ª Região de Saúde": "Macrorregião II",
    "6ª Região de Saúde": "Macrorregião I",
    "7ª Região de Saúde": "Macrorregião II",
    "8ª Região de Saúde": "Macrorregião II",
    "9ª Região de Saúde": "Macrorregião II",
}


@login_required
def report_page_view(request):
    """
    Renderiza a página com o formulário de geração de relatórios.
    """
    return render(request, "manager/report_page.html")


@login_required
def audit_log_view(request):
    """
    Exibe a lista de logs de auditoria do django-auditlog em um template customizado.
    """
    ACTION_MAP = {
        "create": LogEntry.Action.CREATE,
        "update": LogEntry.Action.UPDATE,
        "delete": LogEntry.Action.DELETE,
    }

    queryset = LogEntry.objects.select_related("actor").order_by("-timestamp")

    search_query = request.GET.get("search_query")
    action_type = request.GET.get("action_type")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if search_query:
        queryset = queryset.filter(
            Q(actor__first_name__icontains=search_query)
            | Q(actor__last_name__icontains=search_query)
            | Q(actor__username__icontains=search_query)
            | Q(object_repr__icontains=search_query)
            | Q(changes_text__icontains=search_query)
        )

    if action_type in ACTION_MAP:
        queryset = queryset.filter(action=ACTION_MAP[action_type])

    if start_date:
        queryset = queryset.filter(timestamp__date__gte=start_date)

    if end_date:
        queryset = queryset.filter(timestamp__date__lte=end_date)

    paginator = Paginator(queryset, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if "page" in query_params:
        del query_params["page"]

    context = {
        "page_obj": page_obj,
        "query_params": query_params.urlencode(),
    }
    return render(request, "manager/audit_log.html", context)


def _calculate_age_string(birth_date, gestational_weeks):
    """Calcula a idade cronológica e corrigida e retorna uma string formatada."""
    if not birth_date:
        return "Idade desconhecida"

    today = timezone.now().date()

    chrono_days = (today - birth_date).days
    chrono_weeks = chrono_days // 7
    chrono_remaining_days = chrono_days % 7

    corrected_str = ""
    if gestational_weeks < 37:
        prematurity_days = (40 - gestational_weeks) * 7
        corrected_days = max(0, chrono_days - prematurity_days)
        corrected_weeks = corrected_days // 7
        corrected_str = f" (IC: {corrected_weeks}s)"

    return f"{chrono_weeks}s + {chrono_remaining_days}d{corrected_str}"


@login_required
def manager_dashboard(request):
    """
    Renderiza o template do painel principal do gestor.
    """
    return render(request, "manager/manager-dashboard.html")


@login_required
def dashboard_stats_api(request):
    """
    API view que retorna TODOS os dados dinâmicos para o painel do gestor.
    """
    today = timezone.now().date()
    current_year = today.year

    all_patients = Patient.objects.filter(is_active=True).order_by("-created_at")

    total_ativos_count = all_patients.count()

    children_acompanhamento = []
    children_estaveis = []
    children_alerta = []
    children_critico = []

    estaveis_count = 0
    alerta_count = 0
    critico_count = 0

    for patient in all_patients:
        age_str = _calculate_age_string(patient.date_of_birth, patient.gestational_age_weeks)
        patient_data = {
            "id": patient.id,
            "name": str(patient),
            "age": age_str,
            "status": "Acompanhamento regular",
            "real_status": "estaveis",
        }

        latest_consultation = (
            Record.objects.filter(patient=patient, record_type="consultation")
            .order_by("-date", "-created_at")
            .first()
        )

        sinal_count = 0
        consulta_atrasada = False
        data_atraso_str = ""

        if latest_consultation:
            sinal_count = latest_consultation.warning_signs.filter(is_present=True).count()  # type: ignore

            try:
                consultation_details = latest_consultation.consultation_details  # type: ignore

                if (
                    consultation_details
                    and consultation_details.next_appointment_date
                    and consultation_details.next_appointment_date < today
                ):
                    consulta_atrasada = True
                    data_atraso_str = f"Consulta atrasada desde {consultation_details.next_appointment_date.strftime('%d/%m')}"
            except (ConsultationRecord.DoesNotExist, AttributeError):
                pass

        if sinal_count >= 2:
            critico_count += 1
            patient_data["real_status"] = "critico"
            sign = (
                latest_consultation.warning_signs.filter(is_present=True).first()  # type: ignore
                if latest_consultation
                else None
            )
            patient_data["status"] = (
                f"Crítico: {sign.get_type_display() if sign else 'Múltiplos sinais'}"  # type: ignore
            )
            children_critico.append(patient_data)

        elif sinal_count == 1 or consulta_atrasada:
            alerta_count += 1
            patient_data["real_status"] = "alerta"
            if sinal_count == 1:
                sign = (
                    latest_consultation.warning_signs.filter(is_present=True).first()  # type: ignore
                    if latest_consultation
                    else None
                )
                patient_data["status"] = f"Alerta: {sign.get_type_display() if sign else 'Alerta'}"  # type: ignore
            else:
                patient_data["status"] = data_atraso_str
            children_alerta.append(patient_data)

        else:
            estaveis_count += 1
            patient_data["real_status"] = "estaveis"
            children_estaveis.append(patient_data)

    children_acompanhamento = children_estaveis + children_alerta + children_critico
    children_acompanhamento = children_acompanhamento[:20]

    children_alerta = children_alerta[:20]
    children_critico = children_critico[:20]
    children_estaveis = children_estaveis[:20]

    visits_completed_this_month = Record.objects.filter(
        record_type="consultation", date__year=today.year, date__month=today.month
    ).count()

    pending_visits_count = ConsultationRecord.objects.filter(
        next_appointment_date__year=today.year,
        next_appointment_date__month=today.month,
        next_appointment_date__gte=today,
    ).count()
    overdue_visits_count = ConsultationRecord.objects.filter(
        next_appointment_date__lt=today
    ).count()
    top_professionals = list(
        Record.objects.filter(professional__isnull=False)
        .exclude(professional__exact="")
        .values("professional")
        .annotate(attendances=Count("id"))
        .order_by("-attendances")[:4]
    )

    discharge_chart_data = {}
    patient_content_type = ContentType.objects.get_for_model(Patient)

    for year in range(current_year - 2, current_year + 1):
        monthly_counts = [0] * 12
        discharges = (
            LogEntry.objects.filter(
                content_type=patient_content_type,
                action=LogEntry.Action.UPDATE,
                timestamp__year=year,
                changes_text__icontains='"is_active": [true, false]',
            )
            .annotate(month=TruncMonth("timestamp"))
            .values("month")
            .annotate(count=Count("id"))
            .order_by("month")
        )

        for item in discharges:
            monthly_counts[item["month"].month - 1] = item["count"]
        discharge_chart_data[year] = monthly_counts

    data = {
        "stats": {
            "babies_in_stage_3": total_ativos_count,
            "casos_alerta": alerta_count,
            "casos_criticos": critico_count,
            "casos_estaveis": estaveis_count,
            "visits_completed": visits_completed_this_month,
            "visits_pending": pending_visits_count,
            "overdue_visits": overdue_visits_count,
        },
        "monitored_children": {
            "acompanhamento": children_acompanhamento,
            "alerta": children_alerta,
            "critico": children_critico,
            "estaveis": children_estaveis,
        },
        "top_professionals": top_professionals,
        "discharge_chart": discharge_chart_data,
        "last_update": timezone.now().strftime("hoje às %H:%M"),
    }

    return JsonResponse(data)


def _titlecase_city(name: str) -> str:
    if not name:
        return "Sem Município"
    return " ".join(p.capitalize() for p in name.split())


def create_count_entry():
    return {"total": 0, "alerta": 0, "critico": 0, "estavel": 0}


def api_map_counts(request):
    """
    API que retorna as contagens de pacientes (total, alerta, critico)
    agregadas.
    """
    today = timezone.now().date()

    counts_municipio = defaultdict(create_count_entry)
    counts_microrregiao = defaultdict(create_count_entry)
    counts_macrorregiao = defaultdict(create_count_entry)

    patients = Patient.objects.filter(is_active=True)

    for patient in patients:
        city = patient.address_city

        if not city:
            continue

        city_normalized = normalize_str(patient.address_city)

        status = "estavel"
        sinal_count = 0
        consulta_atrasada = False

        latest_consultation = (
            Record.objects.filter(patient=patient, record_type="consultation")
            .order_by("-date", "-created_at")
            .first()
        )

        if latest_consultation:
            # 4. Busca os sinais e detalhes (N+1 query)
            sinal_count = latest_consultation.warning_signs.filter(is_present=True).count()  # type: ignore

            try:
                consultation_details = latest_consultation.consultation_details  # type: ignore

                if (
                    consultation_details
                    and consultation_details.next_appointment_date
                    and consultation_details.next_appointment_date < today
                ):
                    consulta_atrasada = True
            except (ConsultationRecord.DoesNotExist, AttributeError):
                pass

        if sinal_count >= 2:
            status = "critico"
        elif sinal_count == 1 or consulta_atrasada:
            status = "alerta"
        else:
            status = "estavel"

        counts_municipio[city_normalized]["total"] += 1
        counts_municipio[city_normalized][status] += 1

        microrregiao = MICRORREGIAO_POR_MUNICIPIO.get(city_normalized)
        if microrregiao:
            counts_microrregiao[microrregiao]["total"] += 1
            counts_microrregiao[microrregiao][status] += 1

            macrorregiao = MACRORREGIAO_POR_MICRORREGIAO.get(microrregiao)
            if macrorregiao:
                counts_macrorregiao[macrorregiao]["total"] += 1
                counts_macrorregiao[macrorregiao][status] += 1

    return JsonResponse(
        {
            "municipio": dict(counts_municipio),
            "microrregiao": dict(counts_microrregiao),
            "macrorregiao": dict(counts_macrorregiao),
        }
    )


def create_report_entry():
    """Cria a estrutura de dicionário necessária para agregar dados do relatório."""
    return {
        "ativos": 0,
        "critico": 0,
        "alerta": 0,
        "estavel": 0,
        "finalizados_periodo": 0,
        "sum_gest_age_weeks": 0,
        "sum_birth_weight": Decimal("0.0"),
        "count_for_avg": 0,
        "warning_signs": defaultdict(int),
        "count_ext_preterm": 0,  # < 32 semanas (Muito Prematuro)
        "count_vlbw": 0,  # < 1500g (Very Low Birth Weight)
        "count_elbw": 0,  # < 1000g (Extremely Low Birth Weight)
        "count_ame": 0,  # Aleitamento Materno Exclusivo
    }


def _get_report_data(start_date=None, end_date=None):
    """
    Busca e agrega dados por município.
    VERSÃO OTIMIZADA E ENRIQUECIDA.
    """
    today = timezone.now().date()
    data_by_city = defaultdict(create_report_entry)
    sign_display_map = dict(ClinicalWarningSign.WarningSignType.choices)

    latest_consult_qs = (
        Record.objects.filter(record_type="consultation")
        .order_by("-date", "-created_at")
        .select_related("consultation_details")
        .prefetch_related("warning_signs")
    )

    active_patients = Patient.objects.filter(is_active=True).prefetch_related(
        Prefetch("records", queryset=latest_consult_qs, to_attr="consultation_records")
    )

    print(f"DEBUG: Encontrados {active_patients.count()} pacientes ativos.")

    for patient in active_patients:
        city_normalized = normalize_str(patient.address_city)
        if not city_normalized:
            print(f"DEBUG: Pulando paciente {patient.id} sem cidade.")
            continue

        region_data = data_by_city[city_normalized]
        region_data["ativos"] += 1

        status = "estavel"
        sinal_count = 0
        consulta_atrasada = False
        warning_signs_present_codes = []

        latest_consultation = (
            patient.consultation_records[0] if patient.consultation_records else None
        )

        if latest_consultation:
            try:
                signs_present = [s for s in latest_consultation.warning_signs.all() if s.is_present]
                sinal_count = len(signs_present)
                warning_signs_present_codes = [s.type for s in signs_present]

                consultation_details = latest_consultation.consultation_details
                if (
                    consultation_details
                    and consultation_details.next_appointment_date
                    and consultation_details.next_appointment_date < today
                ):
                    consulta_atrasada = True

                if consultation_details and consultation_details.feeding_type == "exclusive":
                    region_data["count_ame"] += 1

            except (ConsultationRecord.DoesNotExist, AttributeError):
                pass
            except Exception as e:
                print(
                    f"DEBUG: Erro ao processar consulta {latest_consultation.id} para paciente {patient.id}: {e}"
                )

        if sinal_count >= 2:
            status = "critico"
        elif sinal_count == 1 or consulta_atrasada:
            status = "alerta"
        else:
            status = "estavel"

        region_data[status] += 1

        has_demographic = False
        if patient.gestational_age_weeks is not None:
            region_data["sum_gest_age_weeks"] += patient.gestational_age_weeks
            has_demographic = True

            if patient.gestational_age_weeks < 32:
                region_data["count_ext_preterm"] += 1

        if patient.birth_weight is not None:
            region_data["sum_birth_weight"] += patient.birth_weight
            has_demographic = True

            if patient.birth_weight < 1500:
                region_data["count_vlbw"] += 1
            if patient.birth_weight < 1000:
                region_data["count_elbw"] += 1

        if has_demographic:
            region_data["count_for_avg"] += 1

        for sign_code in warning_signs_present_codes:
            sign_name = sign_display_map.get(sign_code, sign_code)
            region_data["warning_signs"][sign_name] += 1

    if start_date and end_date:
        patient_content_type = ContentType.objects.get_for_model(Patient)
        discharge_logs = (
            LogEntry.objects.filter(
                content_type=patient_content_type,
                action=LogEntry.Action.UPDATE,
                changes_text__icontains='"is_active": [true, false]',
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date,
            )
            .values_list("object_pk", flat=True)
            .distinct()
        )

        discharged_patient_ids = [pk for pk in discharge_logs if pk]

        discharged_patients_info = Patient.objects.filter(pk__in=discharged_patient_ids).values(
            "pk", "address_city"
        )

        city_map = {
            str(p["pk"]): normalize_str(p["address_city"])
            for p in discharged_patients_info
            if p["address_city"]
        }

        for patient_id in discharged_patient_ids:
            city_normalized = city_map.get(str(patient_id))
            if city_normalized:
                data_by_city[city_normalized]["finalizados_periodo"] += 1
        print(f"DEBUG: Processados {len(discharged_patient_ids)} pacientes finalizados no período.")
    else:
        print("DEBUG: Sem intervalo de datas, contagem de finalizados será 0.")

    return data_by_city


def _aggregate_by_region(data_by_city, level):
    """Agrega dados de município para micro/macro e calcula métricas finais."""
    if level not in ["municipio", "microrregiao", "macrorregiao"]:
        return []

    aggregated_data = defaultdict(create_report_entry)

    if level == "municipio":
        aggregated_data = data_by_city
    elif level == "microrregiao":
        for city, city_data in data_by_city.items():
            micro_region = MICRORREGIAO_POR_MUNICIPIO.get(city, "Microrregião Desconhecida")
            region_agg = aggregated_data[micro_region]
            for key in [
                "ativos",
                "critico",
                "alerta",
                "estavel",
                "finalizados_periodo",
                "sum_gest_age_weeks",
                "count_for_avg",
                "count_ext_preterm",
                "count_vlbw",
                "count_elbw",
                "count_ame",
            ]:
                region_agg[key] += city_data[key]
            region_agg["sum_birth_weight"] += city_data["sum_birth_weight"]
            for sign, count in city_data["warning_signs"].items():
                region_agg["warning_signs"][sign] += count
    elif level == "macrorregiao":
        for city, city_data in data_by_city.items():
            micro_region = MICRORREGIAO_POR_MUNICIPIO.get(city)
            macro_region = (
                MACRORREGIAO_POR_MICRORREGIAO.get(micro_region, "Macrorregião Desconhecida")
                if micro_region
                else "Macrorregião Desconhecida"
            )
            region_agg = aggregated_data[macro_region]
            for key in [
                "ativos",
                "critico",
                "alerta",
                "estavel",
                "finalizados_periodo",
                "sum_gest_age_weeks",
                "count_for_avg",
                "count_ext_preterm",
                "count_vlbw",
                "count_elbw",
                "count_ame",
            ]:
                region_agg[key] += city_data[key]
            region_agg["sum_birth_weight"] += city_data["sum_birth_weight"]
            for sign, count in city_data["warning_signs"].items():
                region_agg["warning_signs"][sign] += count

    report_list = []
    for region_name, data in aggregated_data.items():
        count_avg = data["count_for_avg"]
        ativos_count = data["ativos"] if data["ativos"] > 0 else 1

        avg_gest_age = (
            round(Decimal(data["sum_gest_age_weeks"]) / Decimal(count_avg), 1)
            if count_avg > 0
            else Decimal("0.0")
        )
        avg_birth_weight = (
            (data["sum_birth_weight"] / Decimal(count_avg)).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
            if count_avg > 0
            else Decimal("0")
        )

        sorted_signs = sorted(data["warning_signs"].items(), key=lambda item: item[1], reverse=True)
        top_signs_formatted = [f"{name} ({count})" for name, count in sorted_signs[:3]]

        report_list.append(
            {
                "regiao": region_name,
                "ativos": data["ativos"],
                "critico": data["critico"],
                "alerta": data["alerta"],
                "estavel": data["estavel"],
                "perc_critico": round(Decimal(data["critico"] * 100) / Decimal(ativos_count), 1),
                "perc_alerta": round(Decimal(data["alerta"] * 100) / Decimal(ativos_count), 1),
                "finalizados_periodo": data["finalizados_periodo"],
                "avg_gest_age": avg_gest_age,
                "avg_birth_weight": avg_birth_weight,
                "top_warning_signs": top_signs_formatted,
                "perc_ame": round(Decimal(data["count_ame"] * 100) / Decimal(ativos_count), 1),
                "perc_ext_preterm": round(
                    Decimal(data["count_ext_preterm"] * 100) / Decimal(ativos_count), 1
                ),
                "perc_vlbw": round(Decimal(data["count_vlbw"] * 100) / Decimal(ativos_count), 1),
                "perc_elbw": round(Decimal(data["count_elbw"] * 100) / Decimal(ativos_count), 1),
            }
        )

    report_list.sort(key=lambda x: x["regiao"])
    return report_list


def _generate_excel(report_data, level, start_date, end_date):
    """Gera um arquivo Excel (.xlsx) com os novos campos."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Relatório Karu - {level.capitalize()}"

    title = f"Relatório Karu - Nível: {level.capitalize()}"
    if start_date and end_date:
        title += (
            f" (Finalizados: {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')})"
        )
    sheet.append([title])

    sheet.merge_cells("A1:O1")
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append([])

    headers = [
        "Região",
        "Ativos",
        "Críticos",
        "% Críticos",
        "Alertas",
        "% Alertas",
        "Estáveis",
        "Finalizados",
        "Média IG (sem)",
        "Média Peso (g)",
        "% AME",
        "% Prem. <32s",
        "% MBP (<1500g)",
        "% EBP (<1000g)",
        "Principais Sinais (Top 3)",
    ]
    sheet.append(headers)
    header_font = Font(bold=True)
    for cell in sheet[3]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in report_data:
        row = [
            item["regiao"],
            item["ativos"],
            item["critico"],
            f"{item['perc_critico']}%",
            item["alerta"],
            f"{item['perc_alerta']}%",
            item["estavel"],
            item["finalizados_periodo"],
            item["avg_gest_age"],
            item["avg_birth_weight"],
            f"{item['perc_ame']}%",
            f"{item['perc_ext_preterm']}%",
            f"{item['perc_vlbw']}%",
            f"{item['perc_elbw']}%",
            ", ".join(item["top_warning_signs"]),
        ]
        sheet.append(row)

        for cell_idx in range(2, 16):
            sheet.cell(row=sheet.max_row, column=cell_idx).alignment = Alignment(
                horizontal="center"
            )

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 12
    sheet.column_dimensions["G"].width = 10
    sheet.column_dimensions["H"].width = 12
    sheet.column_dimensions["I"].width = 15
    sheet.column_dimensions["J"].width = 15
    sheet.column_dimensions["K"].width = 12  # % AME
    sheet.column_dimensions["L"].width = 14  # % Prem.
    sheet.column_dimensions["M"].width = 14  # % MBP
    sheet.column_dimensions["N"].width = 14  # % EBP
    sheet.column_dimensions["O"].width = 60  # Sinais

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_pdf(report_data, level, start_date, end_date):
    """Gera um PDF usando xhtml2pdf."""
    now_local = timezone.localtime(timezone.now())

    context = {
        "report_data": report_data,
        "level": level.capitalize(),
        "start_date": start_date.strftime("%d/%m/%Y") if start_date else None,
        "end_date": end_date.strftime("%d/%m/%Y") if end_date else None,
        "generation_date": now_local.strftime("%d/%m/%Y %H:%M:%S"),
    }
    html_string = render_to_string("manager/report_template.html", context)
    buffer = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer)
    if pisa_status.err:
        print(f"Erro ao gerar PDF: {pisa_status.err}")
        return None
    buffer.seek(0)
    return buffer.getvalue()


@login_required
def generate_report_view(request):
    """View principal para gerar e baixar relatórios."""
    level = request.GET.get("level", "municipio").lower()
    report_format = request.GET.get("format", "excel").lower()
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    start_date, end_date = None, None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        elif start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = timezone.now().date()
    except ValueError:
        return HttpResponse("Formato de data inválido. Use AAAA-MM-DD.", status=400)

    if start_date and end_date and end_date < start_date:
        return HttpResponse("A data final não pode ser anterior à data inicial.", status=400)

    print(
        f"Iniciando geração do relatório: Nível='{level}', Formato='{report_format}', Datas='{start_date}' a '{end_date}'"
    )
    data_by_city = _get_report_data(start_date, end_date)
    report_data = _aggregate_by_region(data_by_city, level)

    if not report_data:
        print(f"WARNING: Nenhum dado encontrado para o nível '{level}'.")
        return HttpResponse(
            f"Nenhum dado encontrado para o nível geográfico '{level}'.", status=404
        )

    file_content = None
    content_type = ""
    filename = f"relatorio_karu_{level}_{timezone.now().strftime('%Y%m%d')}"

    try:
        if report_format == "excel":
            file_content = _generate_excel(report_data, level, start_date, end_date)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename += ".xlsx"
        elif report_format == "pdf":
            file_content = _generate_pdf(report_data, level, start_date, end_date)
            content_type = "application/pdf"
            filename += ".pdf"
        else:
            return HttpResponse("Formato de relatório inválido. Use 'pdf' ou 'excel'.", status=400)

        if file_content is None:
            raise Exception(f"Função de geração para '{report_format}' retornou None.")

    except Exception as e:
        print(f"ERRO CRÍTICO ao gerar relatório '{report_format}': {e}")
        return HttpResponse(
            f"Erro interno ao gerar o relatório {report_format.upper()}.", status=500
        )

    print(f"Relatório gerado com sucesso: {filename}")
    response = HttpResponse(file_content, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def api_report_charts_data(request):
    """
    Endpoint da API que retorna dados agregados para os gráficos
    do painel de relatórios.

    VERSÃO ATUALIZADA: Inclui 6 DESTAQUES REGIONAIS.
    """
    level = request.GET.get("level", "municipio").lower()
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    start_date, end_date = None, None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"error": "Formato de data inválido."}, status=400)

    data_by_city = _get_report_data(start_date, end_date)
    report_data = _aggregate_by_region(data_by_city, level)

    top_critico_regiao = {"regiao": "N/D", "perc": 0.0}
    top_alerta_regiao = {"regiao": "N/D", "perc": 0.0}
    top_risco_prem_regiao = {"regiao": "N/D", "perc": 0.0}
    top_vlbw_regiao = {"regiao": "N/D", "perc": 0.0}
    top_elbw_regiao = {"regiao": "N/D", "perc": 0.0}
    bottom_ame_regiao = {"regiao": "N/D", "perc": 0.0}

    geo_labels = []
    geo_data_ativos = []
    geo_data_criticos = []
    geo_data_alertas = []
    perc_criticos_data = []
    perc_alertas_data = []

    total_ativos = 0
    total_estaveis = 0
    total_alertas = 0
    total_criticos = 0
    total_prem_32 = Decimal("0.0")
    total_vlbw_1500 = Decimal("0.0")
    total_elbw_1000 = Decimal("0.0")

    if report_data:
        report_data_with_ativos = [x for x in report_data if x["ativos"] > 0]

        if report_data_with_ativos:
            top_critico_regiao_data = max(report_data_with_ativos, key=lambda x: x["perc_critico"])
            top_critico_regiao = {
                "regiao": top_critico_regiao_data["regiao"],
                "perc": top_critico_regiao_data["perc_critico"],
            }

            top_alerta_regiao_data = max(report_data_with_ativos, key=lambda x: x["perc_alerta"])
            top_alerta_regiao = {
                "regiao": top_alerta_regiao_data["regiao"],
                "perc": top_alerta_regiao_data["perc_alerta"],
            }

            top_risco_prem_regiao_data = max(
                report_data_with_ativos, key=lambda x: x["perc_ext_preterm"]
            )
            top_risco_prem_regiao = {
                "regiao": top_risco_prem_regiao_data["regiao"],
                "perc": top_risco_prem_regiao_data["perc_ext_preterm"],
            }

            top_vlbw_regiao_data = max(report_data_with_ativos, key=lambda x: x["perc_vlbw"])
            top_vlbw_regiao = {
                "regiao": top_vlbw_regiao_data["regiao"],
                "perc": top_vlbw_regiao_data["perc_vlbw"],
            }

            top_elbw_regiao_data = max(report_data_with_ativos, key=lambda x: x["perc_elbw"])
            top_elbw_regiao = {
                "regiao": top_elbw_regiao_data["regiao"],
                "perc": top_elbw_regiao_data["perc_elbw"],
            }

            bottom_ame_regiao_data = min(report_data_with_ativos, key=lambda x: x["perc_ame"])
            bottom_ame_regiao = {
                "regiao": bottom_ame_regiao_data["regiao"],
                "perc": bottom_ame_regiao_data["perc_ame"],
            }

        for item in report_data:
            geo_labels.append(item["regiao"])
            geo_data_ativos.append(item["ativos"])
            geo_data_criticos.append(item["critico"])
            geo_data_alertas.append(item["alerta"])
            perc_criticos_data.append(item["perc_critico"])
            perc_alertas_data.append(item["perc_alerta"])

            total_ativos += item["ativos"]
            total_estaveis += item["estavel"]
            total_alertas += item["alerta"]
            total_criticos += item["critico"]
            total_prem_32 += item["ativos"] * item["perc_ext_preterm"]
            total_vlbw_1500 += item["ativos"] * item["perc_vlbw"]
            total_elbw_1000 += item["ativos"] * item["perc_elbw"]

    status_data = [total_estaveis, total_alertas, total_criticos]
    if total_ativos > 0 and total_alertas == 0 and total_criticos == 0:
        status_data = [total_estaveis]
    elif total_ativos == 0:
        status_data = [0, 0, 0]

    safe_total_ativos = total_ativos if total_ativos > 0 else 1
    risk_data = [
        round(total_prem_32 / safe_total_ativos, 1),
        round(total_vlbw_1500 / safe_total_ativos, 1),
        round(total_elbw_1000 / safe_total_ativos, 1),
    ]

    data = {
        "geo_distribution": {
            "labels": geo_labels,
            "datasets": [
                {
                    "label": "Críticos",
                    "data": geo_data_criticos,
                    "backgroundColor": "rgba(239, 68, 68, 0.7)",
                },
                {
                    "label": "Alertas",
                    "data": geo_data_alertas,
                    "backgroundColor": "rgba(245, 158, 11, 0.7)",
                },
                {
                    "label": "Total Ativos",
                    "data": geo_data_ativos,
                    "backgroundColor": "rgba(59, 130, 246, 0.7)",
                },
            ],
        },
        "status_distribution": {
            "labels": ["Estáveis", "Alertas", "Críticos"],
            "data": status_data,
        },
        "risk_profile": {
            "labels": ["% Prem. <32s", "% MBP <1500g", "% EBP <1000g"],
            "data": risk_data,
        },
        "proportional_status": {
            "labels": geo_labels,
            "datasets": [
                {
                    "label": "% Críticos",
                    "data": perc_criticos_data,
                    "backgroundColor": "rgba(239, 68, 68, 0.7)",
                },
                {
                    "label": "% Alertas",
                    "data": perc_alertas_data,
                    "backgroundColor": "rgba(245, 158, 11, 0.7)",
                },
            ],
        },
        "summary_kpis": {
            "total_ativos": total_ativos,
            "total_criticos": total_criticos,
            "total_alertas": total_alertas,
            "total_finalizados": sum(item["finalizados_periodo"] for item in report_data)
            if report_data
            else 0,
        },
        "regional_highlights": {
            "top_critico": top_critico_regiao,
            "top_alerta": top_alerta_regiao,
            "top_risco_prem": top_risco_prem_regiao,
            "top_vlbw": top_vlbw_regiao,
            "top_elbw": top_elbw_regiao,
            "bottom_ame": bottom_ame_regiao,
        },
    }

    def decimal_default(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        raise TypeError

    return JsonResponse(data, json_dumps_params={"default": decimal_default})
